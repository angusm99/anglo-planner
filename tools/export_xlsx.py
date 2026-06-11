"""Export job rows from the Google Sheets planner xlsx to data/import.json.

Reads monthly tabs (e.g. MAY-2026) and JOBS IN QUEUE. Data starts at row 4
(matches TARGET_DATA_START_ROW in the sheet's sync script); repeated header
rows inside the data area are skipped by content checks.
"""
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

SOURCE = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(r"C:\Users\angusm\planner_source.xlsx")
OUT = Path(__file__).resolve().parent.parent / "data" / "import.json"

MONTH_TAB = re.compile(r"^(JANUARY|FEBRUARY|MARCH|APRIL|MAY|JUNE|JULY|AUGUST|SEPTEMBER|OCTOBER|NOVEMBER|DECEMBER)-\d{4}$")


def cell_str(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def cell_num(v):
    try:
        return float(v) if v is not None and str(v).strip() != "" else None
    except (TypeError, ValueError):
        return None


def cell_date(v):
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = cell_str(v)
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    return m.group(0) if m else None


def main():
    wb = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)
    tabs = [n for n in wb.sheetnames if MONTH_TAB.match(n)] + ["JOBS IN QUEUE"]
    records = []

    for tab in tabs:
        ws = wb[tab]
        for row in ws.iter_rows(min_row=4, max_col=22, values_only=True):
            task_no = cell_str(row[0])
            biz_ref = cell_str(row[3])
            customer = cell_str(row[4])
            if not biz_ref and not task_no:
                continue
            # Skip repeated in-sheet header rows
            if "REF" in task_no.upper() or "BIZMAN" in biz_ref.upper() or not customer:
                continue
            records.append({
                "task_no": task_no,
                "biz_ref": biz_ref,
                "install_date": cell_date(row[1]),
                "send_to_dash": cell_str(row[2]),
                "customer": customer,
                "colour": cell_str(row[5]),
                "qty_windows": cell_num(row[6]),
                "qty_hinged": cell_num(row[7]),
                "qty_folding": cell_num(row[8]),
                "qty_palace": cell_num(row[9]),
                "qty_specials": cell_num(row[10]),
                "qty_elite": cell_num(row[11]),
                "glasslist": 1 if str(row[12]).strip().upper() == "TRUE" else 0,
                "s1": cell_str(row[14]).upper(),
                "s2": cell_str(row[15]).upper(),
                "s3": cell_str(row[16]).upper(),
                "s4": cell_str(row[17]).upper(),
                "s5": cell_str(row[18]).upper(),
                "s6": cell_str(row[19]).upper(),
                "s7": cell_str(row[20]).upper(),
                "job_status": cell_str(row[21]).upper(),
                "source_tab": tab,
            })
    wb.close()

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(records, indent=1), encoding="utf-8")
    print(f"Exported {len(records)} rows from {len(tabs)} tabs -> {OUT}")


if __name__ == "__main__":
    main()
